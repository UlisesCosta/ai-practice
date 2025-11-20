
import os
import csv
from openpyxl import load_workbook

def xlsx_to_csv_folder(xlsx_path: str):
    # Pedir nombre de la carpeta
    folder_name = input("Nombre de la carpeta destino: ").strip()
    if not folder_name:
        raise ValueError("El nombre de carpeta no puede estar vacío.")

    # Crear carpeta si no existe
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    # Cargar workbook sin evaluar fórmulas
    wb = load_workbook(xlsx_path, data_only=False)

    print(f"Convirtiendo '{xlsx_path}'...")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        csv_file_path = os.path.join(folder_name, f"{sheet}.csv")

        with open(csv_file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            for row in ws.iter_rows(values_only=False):
                # Convertir todas las celdas a texto puro
                parsed_row = []
                for cell in row:
                    if cell.value is None:
                        parsed_row.append("")  # dejar vacío
                    else:
                        parsed_row.append(str(cell.value))
                writer.writerow(parsed_row)

        print(f"✔ Hoja '{sheet}' → {csv_file_path}")

    print("\nConversión terminada con éxito.")


if __name__ == "__main__":
    archivo = input("Ruta del archivo .xlsx: ").strip()
    if not archivo.endswith(".xlsx"):
        raise ValueError("Debe proporcionar un archivo .xlsx válido.")
    xlsx_to_csv_folder(archivo)


